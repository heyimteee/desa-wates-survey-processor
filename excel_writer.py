#!/usr/bin/env python3
"""Template-based Excel output writer."""

import os
import re
import datetime

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from config import (
    TEXT_COLUMNS_INDIVIDU, DATE_COLUMNS_INDIVIDU,
    INT_COLUMNS_INDIVIDU, NUM_COLUMNS_INDIVIDU,
    TEXT_COLUMNS_KELUARGA, NUM_COLUMNS_KELUARGA,
)
from utils import auto_fit_columns


def write_individu_output(records, template_path, output_path, logger):
    """Write INDIVIDU records to Excel with proper column types."""
    wb_tpl = openpyxl.load_workbook(template_path)
    ws_tpl = wb_tpl[wb_tpl.sheetnames[0]]
    headers = []
    for c in range(1, ws_tpl.max_column + 1):
        h = ws_tpl.cell(1, c).value
        headers.append(str(h).strip() if h else "")
    wb_tpl.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA INDIVIDU"

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = h
        cell.font = Font(bold=True, name="Calibri", size=11)

    for row_idx, rec in enumerate(records, 2):
        for col_idx, h in enumerate(headers, 1):
            val = rec.get(h)
            cell = ws.cell(row_idx, col_idx)

            if h == "Action":
                if val:
                    cell.value = str(val)
                continue
            elif h.startswith("_"):
                continue

            if h in TEXT_COLUMNS_INDIVIDU:
                if val is not None and val != "":
                    cell.value = str(val)
                    cell.number_format = "@"
                else:
                    cell.value = "-"
                    cell.number_format = "@"
            elif h in DATE_COLUMNS_INDIVIDU:
                if val is not None and isinstance(val, (datetime.date, datetime.datetime)):
                    cell.value = val
                    cell.number_format = "dd-mm-yyyy"
                else:
                    cell.value = "-"
            elif h in INT_COLUMNS_INDIVIDU:
                if val is not None and val != "-":
                    cell.value = int(val) if not isinstance(val, int) else val
                else:
                    cell.value = val if val is not None else "-"
            elif h in NUM_COLUMNS_INDIVIDU:
                if val is not None and val != "-":
                    cell.value = int(val) if isinstance(val, (int, float)) else val
                    cell.number_format = "#,##0"
                else:
                    cell.value = "-"
            else:
                if val is not None and val != "":
                    cell.value = str(val)
                else:
                    cell.value = "-"

    auto_fit_columns(ws, headers, ws.max_row)
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.log("OUTPUT", "INFO",
               f"Wrote {len(records)} rows to {os.path.basename(output_path)}")


def write_keluarga_output(records, template_path, output_path, logger):
    """Write KELUARGA records to Excel with proper column types."""
    wb_tpl = openpyxl.load_workbook(template_path)
    ws_tpl = wb_tpl[wb_tpl.sheetnames[0]]
    headers = []
    for c in range(1, ws_tpl.max_column + 1):
        h = ws_tpl.cell(1, c).value
        headers.append(str(h).strip() if h else "")
    wb_tpl.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA KELUARGA"

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(1, col_idx)
        cell.value = h
        cell.font = Font(name="Calibri", size=11)

    int_pattern = re.compile(r"JARAK|WAKTU|BIAYA", re.IGNORECASE)

    for row_idx, rec in enumerate(records, 2):
        for col_idx, h in enumerate(headers, 1):
            val = rec.get(h)
            cell = ws.cell(row_idx, col_idx)

            if h in TEXT_COLUMNS_KELUARGA:
                if val is not None and val != "" and val != "-":
                    cell.value = str(val)
                    cell.number_format = "@"
                else:
                    cell.value = "-"
                    cell.number_format = "@"
            elif h in NUM_COLUMNS_KELUARGA:
                if val is not None and val != "-" and not isinstance(val, str):
                    cell.value = val
                    if "PENGELUARAN" in h:
                        cell.number_format = "#,##0"
                elif isinstance(val, str) and val != "-":
                    cell.value = val
                else:
                    cell.value = val if val is not None else "-"
            elif int_pattern.search(h):
                if isinstance(val, (int, float)):
                    cell.value = round(val, 2)
                    cell.number_format = "0.00"
                elif val is not None and val != "-":
                    cell.value = val
                else:
                    cell.value = val if val is not None else "-"
            else:
                if val is not None and val != "":
                    cell.value = str(val) if not isinstance(val, (int, float)) else val
                else:
                    cell.value = "-"

    auto_fit_columns(ws, headers, ws.max_row)
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.log("OUTPUT", "INFO",
               f"Wrote {len(records)} rows to {os.path.basename(output_path)}")
