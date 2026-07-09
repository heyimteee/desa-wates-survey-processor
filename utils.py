#!/usr/bin/env python3
"""Formatter, normalizer, and general utility functions."""

import re
import datetime

import openpyxl
from openpyxl.utils import get_column_letter

from config import TODAY, PENDIDIKAN_YEARS


def auto_fit_columns(ws, headers, max_row, min_width=8, max_width=50, padding=2):
    """Auto-size column widths based on content.

    Scans all cells in each column and sets width to fit the longest value,
    capped at max_width to avoid excessively wide columns.
    """
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header)) if header else 0
        for row in range(2, max_row + 1):
            val = ws.cell(row, col_idx).value
            if val is not None:
                s = str(val)
                line_len = max(len(line) for line in s.split("\n"))
                if line_len > max_len:
                    max_len = line_len
        width = min(max(max_len + padding, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def to_upper(val):
    """Convert a value to uppercase string, preserving None/empty."""
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None
    return s.upper()


def to_text(val):
    """Convert a value to stripped string, preserving None/empty."""
    if val is None:
        return None
    s = str(val).strip()
    if s == "":
        return None
    return s


def clean_id(val):
    """Clean an ID value (NIK, KK) from Excel numeric artifacts."""
    if val is None:
        return ""
    if isinstance(val, float):
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    s = str(val).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def to_int(val):
    """Convert a value to int, preserving None/empty."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s == "" or s == "-":
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def clean_health_value(val):
    """Clean a health access survey value for storage.

    Numeric values (1.0, 2.0) -> int. Text values kept as-is. None -> 0.
    """
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s == "":
        return 0
    return s


def normalize_jamkes(val):
    """Normalize Jamkes (health insurance) values."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s == "PESERTA":
        return "PESERTA"
    if "BUKAN" in s:
        return "BUKAN PESERTA"
    return s if s else None


def normalize_yes_no(val):
    """Normalize Ya/Tidak values."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s == "YA":
        return "YA"
    if s == "TIDAK":
        return "TIDAK"
    return s if s else None


def normalize_jk(val):
    """Normalize Jenis Kelamin values."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ("L", "P"):
        return s
    if "LAKI" in s:
        return "L"
    if "PEREMPUAN" in s or "WANITA" in s:
        return "P"
    return s if s else None


def parse_date(val):
    """Parse a date value from BIP (dd-mm-yyyy) or Excel date."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    s = str(val).strip()
    if s == "" or s == "-":
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def calculate_age(birth_date):
    """Calculate age from birth date to today."""
    if birth_date is None:
        return None
    age = TODAY.year - birth_date.year
    if (TODAY.month, TODAY.day) < (birth_date.month, birth_date.day):
        age -= 1
    if age < 0:
        age = 0
    return age


def dash_if_empty(val):
    """Return dash if value is None or empty, otherwise the value."""
    if val is None:
        return "-"
    s = str(val).strip()
    if s == "":
        return "-"
    return val


def format_currency(val):
    """Format a numeric value as integer rupiah (strip decimals)."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s == "" or s == "-":
        return None
    s = s.replace(".", "").replace(",", "")
    try:
        return int(float(s))
    except ValueError:
        return None


def clean_and_validate_phone(val):
    """Clean and validate phone number values.

    Returns (result, note) where result is the cleaned value or '-',
    and note is a validation message or None.

    Rules:
      - None/empty/sentinel (0,1) -> '-', no note
      - +62 prefix -> convert to 0 prefix, note
      - 62 prefix (no +) -> convert to 0 prefix, note
      - Text/letters in value -> '-', note
      - Pure punctuation -> '-', note
      - Single digit -> '-', note
      - 2-6 digits -> '-', note (too short)
      - Commas/spaces stripped from digit sequences
      - Everything else -> digits-only string, no note
    """
    if val is None:
        return "-", None

    if isinstance(val, float):
        if val <= 1:
            return "-", None
        if val == int(val):
            val = int(val)

    if isinstance(val, int):
        if val <= 1:
            return "-", None
        val = str(val)

    s = str(val).strip()
    if s == "" or s in ("0", "1", "0.0", "1.0"):
        return "-", None

    if re.search(r"[a-zA-Z]", s):
        return "-", "mengandung teks"

    cleaned = re.sub(r"\.0+$", "", s)
    cleaned = re.sub(r"[^\d+]", "", cleaned)

    if not cleaned:
        return "-", "hanya tanda baca"

    if cleaned.startswith("+62") and cleaned[3:].isdigit():
        return "0" + cleaned[3:], "dikonversi dari +62 ke 08"

    if cleaned.startswith("62") and cleaned[2:].isdigit():
        return "0" + cleaned[2:], "dikonversi dari 62 ke 08"

    if cleaned.isdigit() and len(cleaned) == 1:
        return "-", "hanya 1 digit"

    if cleaned.isdigit() and len(cleaned) <= 6:
        return "-", "terlalu pendek"

    if cleaned.isdigit() and len(set(cleaned)) <= 2 and len(cleaned) >= 7:
        return cleaned, "pola mencurigakan (digit berulang)"

    if cleaned.isdigit() and not cleaned.startswith("08") and len(cleaned) >= 10:
        return cleaned, "tidak diawali 08"

    return cleaned, None


def clean_email(val):
    """Clean email values, removing Google Forms sentinels."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if val <= 1:
            return None
        return str(int(val))
    s = str(val).strip()
    if s == "" or s in ("0", "1", "0.0", "1.0"):
        return None
    if "@" not in s:
        return None
    return s


def clean_income(val):
    """Clean income values, treating 0/1 sentinels as empty."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if val <= 1:
            return None
        return int(val)
    s = str(val).strip()
    if s == "" or s == "-":
        return None
    s = s.replace(".", "").replace(",", "")
    try:
        n = int(float(s))
        if n <= 1:
            return None
        return n
    except ValueError:
        return None


def is_sentinel(val):
    """Check if a value is a Google Forms sentinel (0 or 1)."""
    if isinstance(val, (int, float)):
        return val in (0, 1, 0.0, 1.0)
    if isinstance(val, str):
        return val.strip() in ("0", "1", "0.0", "1.0")
    return False


def normalize_pendidikan(val):
    """Normalize pendidikan values to template format (spaces around /)."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s == "" or s == "-":
        return None
    s = re.sub(r"\s*/\s*", " / ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def name_key(n):
    """Normalize a name for case-insensitive comparison."""
    if n is None:
        return ""
    return re.sub(r"\s+", " ", str(n).strip()).lower()


def is_valid_nik(nik):
    """Return True if NIK is exactly 16 digits."""
    if not nik:
        return False
    s = str(nik).strip()
    return len(s) == 16 and s.isdigit()


def build_bip_name_index(bip_all):
    """Index BIP records by normalized name.

    Returns:
        dict: name_key -> list of BIP record dicts.
    """
    idx = {}
    for rec in bip_all:
        key = name_key(rec.get("NAMA LENGKAP", ""))
        if key:
            idx.setdefault(key, []).append(rec)
    return idx


def repair_nik_by_name(nik_raw, name, rt, rw, bip_by_name, logger):
    """Attempt to repair an invalid NIK by name matching scoped to RT/RW.

    Returns (repaired_nik, note) or (None, None) if repair failed.
    """
    if is_valid_nik(nik_raw):
        return nik_raw, None

    if not name:
        return None, None

    candidates = bip_by_name.get(name_key(name), [])
    if not candidates:
        return None, None

    rt_str = str(rt).strip().lstrip("0") if rt else None
    rw_str = str(rw).strip().lstrip("0") if rw else None

    scoped = []
    for c in candidates:
        c_rt = str(c.get("NO RT", "")).strip().lstrip("0")
        c_rw = str(c.get("NO RW", "")).strip().lstrip("0")
        if rt_str is None or c_rt == rt_str:
            if rw_str is None or c_rw == rw_str:
                scoped.append(c)

    if len(scoped) == 1:
        new_nik = clean_id(scoped[0].get("NIK"))
        if is_valid_nik(new_nik) and new_nik != nik_raw:
            return new_nik, "diperbaiki dari '{}'".format(nik_raw)

    if len(scoped) > 1:
        return None, "nama '{}' cocok dengan {} data BIP di RT {}/RW {}, ambigu".format(
            name, len(scoped), rt or "?", rw or "?")

    if not scoped and rt is not None:
        full_candidates = candidates
        if len(full_candidates) == 1:
            new_nik = clean_id(full_candidates[0].get("NIK"))
            if is_valid_nik(new_nik) and new_nik != nik_raw:
                return new_nik, "diperbaiki dari '{}' (luar RT/RW)".format(nik_raw)

    return None, "nama tidak ditemukan di BIP RT {}/RW {}".format(
        rt or "?", rw or "?")


def classify_nik_duplicate(nik, name_a, name_b, bip_by_nik):
    """Determine which name a duplicated NIK belongs to in BIP.

    Returns: 'same', 'a_owns', 'b_owns', or 'neither'.
    """
    bip_entry = bip_by_nik.get(nik)
    if not bip_entry:
        return "neither"

    bip_name = (to_upper(bip_entry.get("NAMA LENGKAP")) or "").strip()
    a = (name_a or "").strip().upper()
    b = (name_b or "").strip().upper()

    if bip_name == a == b:
        return "same"
    if bip_name == a:
        return "a_owns"
    if bip_name == b:
        return "b_owns"
    return "neither"


def detect_health_anomaly(val):
    """Detect contradictory or multi-select health values.

    Returns (is_anomaly, anomaly_type, message).
    """
    if val is None or isinstance(val, (int, float)):
        return False, "", ""

    s = str(val).strip()
    if "," not in s:
        return False, "", ""

    parts = [p.strip() for p in s.split(",") if p.strip()]
    if len(parts) < 2:
        return False, "", ""

    has_text = False
    has_numeric = False
    numeric_count = 0

    for p in parts:
        if p.isdigit() or (p.replace(".", "").isdigit()):
            has_numeric = True
            numeric_count += 1
        else:
            has_text = True

    if has_text and has_numeric:
        return True, "mixed", "kontradiksi: '{}' (teks + angka)".format(s)

    if has_numeric and numeric_count > 1:
        return True, "multi_numeric", "multi-select: '{}' ({} angka)".format(
            s, numeric_count)

    return False, "", ""


def derive_pendidikan_years(pendidikan):
    """Derive years of education from pendidikan level."""
    if not pendidikan:
        return None
    p = pendidikan.upper().strip()
    if p in PENDIDIKAN_YEARS:
        return PENDIDIKAN_YEARS[p]
    for key, years in PENDIDIKAN_YEARS.items():
        if key in p or p in key:
            return years
    if "TIDAK" in p or "BELUM" in p:
        return 0
    if "SD" in p:
        return 6
    if "SLTP" in p or "SMP" in p:
        return 9
    if "SLTA" in p or "SMA" in p:
        return 12
    if "DIPLOMA" in p:
        return 13
    if "S-1" in p or "S1" in p or "STRATA I" in p:
        return 16
    if "S-2" in p or "S2" in p or "STRATA II" in p:
        return 18
    if "S-3" in p or "S3" in p or "STRATA III" in p:
        return 21
    return None


def repair_nik_kk_from_bip(kk, bip_by_kk):
    """Find the correct NIK Kepala Keluarga from BIP for a given KK.

    Looks up the KK in BIP and finds the family member whose SHDK
    indicates they are the head of household, returning their NIK.

    Returns:
        (repaired_nik, note) or (None, note) if repair failed.
    """
    members = bip_by_kk.get(str(kk), [])
    if not members:
        return None, "KK tidak ditemukan di BIP"

    for m in members:
        shdk = str(m.get("SHDK", "")).strip().upper()
        if "KEPALA" in shdk:
            nik = clean_id(m.get("NIK"))
            if is_valid_nik(nik):
                return nik, "diperbaiki dari BIP (data kepala keluarga)"

    return None, "tidak ada kepala keluarga di BIP untuk KK ini"


def normalize_bantuan(val):
    """Normalize bantuan value, handling ambiguous YA,TIDAK responses."""
    s = str(val).strip().upper() if val else ""
    if not s or s in ("-", ""):
        return "TIDAK"
    if "YA" in s:
        return "YA"
    if "TIDAK" in s:
        return "TIDAK"
    return s
