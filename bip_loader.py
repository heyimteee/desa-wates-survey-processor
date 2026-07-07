#!/usr/bin/env python3
"""BIP (civil registry) file loader."""

import os
import glob

import openpyxl

from config import BIP_HEADERS
from utils import clean_id
from logger import IssueLogger


def load_bip_pool(bip_dir):
    """Load all BIP Excel files into indexed dictionaries.

    Scans every .xlsx file in bip_dir, auto-detects the header row
    (some files start at row 1, others at row 3), and indexes
    records by NIK and by KK.

    Returns:
        Tuple of (bip_by_nik, bip_by_kk, bip_all, logger).
    """
    logger = IssueLogger()
    bip_by_nik = {}
    bip_by_kk = {}
    bip_all = []

    bip_files = sorted(
        f for f in glob.glob(os.path.join(bip_dir, "*.xlsx"))
        if not os.path.basename(f).startswith("~$")
    )
    if not bip_files:
        logger.log("BIP_LOAD", "WARNING", f"No BIP files found in {bip_dir}")
        return bip_by_nik, bip_by_kk, bip_all, logger

    for fpath in bip_files:
        fname = os.path.basename(fpath)
        print(f"      Loading: {fname}...", end="", flush=True)
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True)
        except Exception as e:
            logger.log("BIP_LOAD", "WARNING", f"Cannot read {fname}: {e}")
            continue

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = None
            for r in range(1, min(6, ws.max_row + 1)):
                val = ws.cell(r, 1).value
                if val is not None and str(val).strip().upper() == "NO":
                    header_row = r
                    break

            if header_row is None:
                logger.log("BIP_LOAD", "WARNING",
                           f"Could not detect header row in {fname}/{sheet_name}, skipping")
                continue

            headers = []
            for c in range(1, ws.max_column + 1):
                h = ws.cell(header_row, c).value
                headers.append(str(h).strip() if h else "")

            col_map = {}
            for idx, h in enumerate(headers):
                if h in BIP_HEADERS:
                    col_map[h] = idx

            for row in range(header_row + 1, ws.max_row + 1):
                record = {}
                for bip_col, col_idx in col_map.items():
                    record[bip_col] = ws.cell(row, col_idx + 1).value

                nik = record.get("NIK")
                if nik is None or str(nik).strip() == "":
                    continue

                nik_str = clean_id(nik)
                record["NIK"] = nik_str
                record["_source_file"] = fname
                record["_source_sheet"] = sheet_name

                kk = record.get("NO KK")
                if kk is not None:
                    record["NO KK"] = clean_id(kk)
                else:
                    record["NO KK"] = ""

                if nik_str in bip_by_nik:
                    logger.log("BIP_DUPLICATE", "INFO",
                               f"NIK {nik_str} appears in multiple BIP entries; "
                               f"keeping first from {bip_by_nik[nik_str]['_source_file']}",
                               nik=nik_str,
                               nama=record.get("NAMA LENGKAP", ""),
                               kk=record.get("NO KK", ""))
                    continue

                bip_by_nik[nik_str] = record
                bip_all.append(record)

                kk_str = record["NO KK"]
                if kk_str:
                    if kk_str not in bip_by_kk:
                        bip_by_kk[kk_str] = []
                    bip_by_kk[kk_str].append(record)

        wb.close()
        print(f" done", flush=True)

    logger.log("BIP_LOAD", "INFO",
               f"Loaded {len(bip_all)} BIP records, "
               f"{len(bip_by_nik)} unique NIKs, {len(bip_by_kk)} unique KKs")
    return bip_by_nik, bip_by_kk, bip_all, logger
