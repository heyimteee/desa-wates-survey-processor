#!/usr/bin/env python3
"""Survey file discovery and loading."""

import os
import re

import openpyxl


def parse_survey_filename(filename):
    """Parse location info from a survey filename.

    Extracts data_type, RT, RW, and Dusun from filenames like:
    - PENDATAAN DATA MIKRO INDIVIDU RT 01 RW 01 DUSUN SIDOMULYO (Jawaban).xlsx
    - Pendataan Data Mikro Keluarga RT 03 RW 05 Dusun Wates ...xlsx

    Returns:
        Dict with keys: data_type, rt, rw, dusun.
    """
    result = {"data_type": None, "rt": None, "rw": None, "dusun": None}
    s = filename.upper()

    if "INDIVIDU" in s:
        result["data_type"] = "INDIVIDU"
    elif "KELUARGA" in s:
        result["data_type"] = "KELUARGA"

    rt_match = re.search(r"RT\s*(\d+)", s)
    if rt_match:
        result["rt"] = rt_match.group(1).lstrip("0") or "0"

    rw_match = re.search(r"RW\s*(\d+)", s)
    if rw_match:
        result["rw"] = rw_match.group(1).lstrip("0") or "0"

    dusun_match = re.search(r"DUSUN\s+([A-Z]+)", s)
    if dusun_match:
        result["dusun"] = dusun_match.group(1)

    return result


def format_rt_range(rt_numbers):
    """Format a list of RT numbers into a compact range string.

    Examples:
        [1, 2, 3, 4, 5] -> "01-05"
        [1, 3, 5] -> "01,03,05"
    """
    if not rt_numbers:
        return "UNKNOWN"
    rts = sorted(set(rt_numbers))
    if len(rts) == 1:
        return f"{rts[0]:02d}"

    runs = []
    start = rts[0]
    end = rts[0]
    for n in rts[1:]:
        if n == end + 1:
            end = n
        else:
            runs.append((start, end))
            start = n
            end = n
    runs.append((start, end))

    parts = []
    for s, e in runs:
        if s == e:
            parts.append(f"{s:02d}")
        else:
            parts.append(f"{s:02d}-{e:02d}")
    return ",".join(parts)


def group_survey_files(filepaths):
    """Group survey files by (data_type, rw, dusun) for output naming.

    Returns:
        Dict mapping (data_type, rw, dusun) -> list of (filepath, parsed_info).
    """
    groups = {}
    for fpath in filepaths:
        fname = os.path.basename(fpath)
        info = parse_survey_filename(fname)
        key = (info["data_type"], info["rw"], info["dusun"])
        if key not in groups:
            groups[key] = []
        groups[key].append((fpath, info))
    return groups


def build_output_name(data_type, rt_numbers, rw, dusun):
    """Build standardized output filename.

    Format: DATA_{type}_RT_{rts}_RW_{rw}_DUSUN_{dusun}_DESA_WATES.xlsx
    """
    rt_str = format_rt_range(rt_numbers)
    rw_str = f"{int(rw):02d}" if rw and rw.isdigit() else (rw or "UNKNOWN")
    dusun_str = (dusun or "UNKNOWN").upper()
    type_str = (data_type or "UNKNOWN").upper()
    return f"DATA_{type_str}_RT_{rt_str}_RW_{rw_str}_DUSUN_{dusun_str}_DESA_WATES.xlsx"


def find_survey_files(bahan_dir, survey_type):
    """Find survey .xlsx files in BAHAN subfolders matching survey_type."""
    pattern_upper = survey_type.upper()
    results = []
    for root, dirs, files in os.walk(bahan_dir):
        for f in files:
            if f.startswith(".") or f.startswith("~$"):
                continue
            if not f.endswith(".xlsx"):
                continue
            if pattern_upper in f.upper():
                results.append(os.path.join(root, f))
    return sorted(results)


def load_individu_survey(filepath):
    """Load an INDIVIDU survey Excel file (Google Forms export).

    Returns:
        List of survey record dicts.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        headers.append(str(h).strip() if h else "")

    records = []
    for row in range(2, ws.max_row + 1):
        record = {}
        for c, h in enumerate(headers):
            record[h] = ws.cell(row, c + 1).value
        if all(v is None or str(v).strip() == "" for v in record.values()):
            continue
        record["_source_file"] = os.path.basename(filepath)
        records.append(record)

    wb.close()
    return records


def load_keluarga_survey(filepath):
    """Load a KELUARGA survey Excel file (Google Forms export).

    Returns:
        List of survey record dicts (one per household).
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = []
    for c in range(1, ws.max_column + 1):
        h = ws.cell(1, c).value
        headers.append(str(h).strip() if h else "")

    records = []
    for row in range(2, ws.max_row + 1):
        record = {}
        for c, h in enumerate(headers):
            record[h] = ws.cell(row, c + 1).value
        if all(v is None or str(v).strip() == "" for v in record.values()):
            continue
        record["_source_file"] = os.path.basename(filepath)
        records.append(record)

    wb.close()
    return records
