#!/usr/bin/env python3
"""Data quality issue logger."""

import openpyxl
from openpyxl.styles import Font

from utils import auto_fit_columns


class IssueLogger:
    """Collects data quality warnings for the DATA_ISSUES output sheet."""

    def __init__(self):
        self.issues = []
        self.validations = []

    def log(self, category, severity, description, nik="", nama="", kk=""):
        """Record a data issue.

        Args:
            category: Type of issue (e.g. 'DUPLICATE_NIK', 'BIP_NOT_FOUND').
            severity: 'WARNING' or 'INFO'.
            description: Human-readable explanation.
            nik: Related NIK if applicable.
            nama: Related name if applicable.
            kk: Related KK if applicable.
        """
        self.issues.append({
            "Category": category,
            "Severity": severity,
            "NIK": nik,
            "Nama": nama,
            "KK": kk,
            "Description": description,
        })

    def log_validation(self, row_no, kk, nik, nama, category, notes):
        """Record a per-row validation entry for the DATA_VALIDASI sheet."""
        self.validations.append({
            "No": row_no,
            "KK": kk,
            "NIK": nik,
            "Nama": nama,
            "Kategori": category,
            "Catatan": notes,
        })

    def write_to_excel(self, filepath):
        """Write all collected issues and validations to an Excel file."""
        wb = openpyxl.Workbook()

        ws = wb.active
        ws.title = "DATA ISSUES"
        headers = ["Category", "Severity", "NIK", "Nama", "KK", "Description"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.value = h
            cell.font = Font(bold=True)

        for row_idx, issue in enumerate(self.issues, 2):
            ws.cell(row_idx, 1).value = issue["Category"]
            ws.cell(row_idx, 2).value = issue["Severity"]
            ws.cell(row_idx, 3).value = issue["NIK"]
            ws.cell(row_idx, 4).value = issue["Nama"]
            ws.cell(row_idx, 5).value = issue["KK"]
            ws.cell(row_idx, 6).value = issue["Description"]

        auto_fit_columns(ws, headers, ws.max_row, min_width=10, max_width=80)

        if self.validations:
            ws2 = wb.create_sheet("DATA VALIDASI")
            vheaders = ["No", "KK", "NIK", "Nama", "Kategori", "Catatan"]
            for col, h in enumerate(vheaders, 1):
                cell = ws2.cell(1, col)
                cell.value = h
                cell.font = Font(bold=True)

            for row_idx, v in enumerate(self.validations, 2):
                ws2.cell(row_idx, 1).value = v["No"]
                ws2.cell(row_idx, 2).value = v["KK"]
                ws2.cell(row_idx, 3).value = v["NIK"]
                ws2.cell(row_idx, 4).value = v["Nama"]
                ws2.cell(row_idx, 5).value = v["Kategori"]
                ws2.cell(row_idx, 6).value = v["Catatan"]

            auto_fit_columns(ws2, vheaders, len(self.validations) + 1,
                             min_width=10, max_width=80)

        wb.save(filepath)
