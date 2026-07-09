"""Tests for survey_loader.py."""

import os

import openpyxl
import pytest

from survey_loader import (
    parse_survey_filename,
    format_rt_range,
    build_output_name,
    find_survey_files,
    load_individu_survey,
    load_keluarga_survey,
)


class TestParseSurveyFilename:
    def test_individu(self):
        result = parse_survey_filename(
            "PENDATAAN DATA MIKRO INDIVIDU RT 05 RW 01 DUSUN SIDOMULYO (Jawaban).xlsx")
        assert result["data_type"] == "INDIVIDU"
        assert result["rt"] == "5"
        assert result["rw"] == "1"
        assert result["dusun"] == "SIDOMULYO"

    def test_keluarga(self):
        result = parse_survey_filename(
            "Pendataan Data Mikro Keluarga RT 03 RW 05 Dusun Wates Desa Wates Tahun 2025 (Jawaban).xlsx")
        assert result["data_type"] == "KELUARGA"
        assert result["rt"] == "3"
        assert result["rw"] == "5"
        assert result["dusun"] == "WATES"

    def test_leading_zero_rt(self):
        result = parse_survey_filename("...RT 01 RW 03...")
        assert result["rt"] == "1"

    def test_unknown_type(self):
        result = parse_survey_filename("some random file.xlsx")
        assert result["data_type"] is None


class TestFormatRtRange:
    def test_single(self):
        assert format_rt_range([3]) == "03"

    def test_contiguous(self):
        assert format_rt_range([1, 2, 3, 4, 5]) == "01-05"

    def test_non_contiguous(self):
        assert format_rt_range([1, 3, 5]) == "01,03,05"

    def test_mixed(self):
        assert format_rt_range([1, 3, 4, 5]) == "01,03-05"

    def test_empty(self):
        assert format_rt_range([]) == "UNKNOWN"


class TestBuildOutputName:
    def test_full(self):
        name = build_output_name("INDIVIDU", [5], "1", "SIDOMULYO")
        assert name == "DATA_INDIVIDU_RT_05_RW_01_DUSUN_SIDOMULYO_DESA_WATES.xlsx"

    def test_keluarga(self):
        name = build_output_name("KELUARGA", [5], "1", "SIDOMULYO")
        assert "DATA_KELUARGA" in name
        assert "RW_01" in name

    def test_unknown_dusun(self):
        name = build_output_name("INDIVIDU", [5], "1", None)
        assert "DUSUN_UNKNOWN" in name


class TestFindSurveyFiles:
    def test_finds_individu(self, temp_dir):
        fpath = os.path.join(temp_dir, "test INDIVIDU RT 01.xlsx")
        open(fpath, "w").close()
        results = find_survey_files(temp_dir, "INDIVIDU")
        assert len(results) == 1

    def test_finds_keluarga(self, temp_dir):
        fpath = os.path.join(temp_dir, "test Keluarga RT 01.xlsx")
        open(fpath, "w").close()
        results = find_survey_files(temp_dir, "Keluarga")
        assert len(results) == 1

    def test_skips_temp_files(self, temp_dir):
        fpath = os.path.join(temp_dir, "~$temp INDIVIDU.xlsx")
        open(fpath, "w").close()
        results = find_survey_files(temp_dir, "INDIVIDU")
        assert len(results) == 0

    def test_subfolder_recursion(self, temp_dir):
        sub = os.path.join(temp_dir, "sub")
        os.makedirs(sub, exist_ok=True)
        fpath = os.path.join(sub, "test INDIVIDU.xlsx")
        open(fpath, "w").close()
        results = find_survey_files(temp_dir, "INDIVIDU")
        assert len(results) == 1


class TestLoadSurvey:
    def test_load_individu(self, temp_survey_indiv_xlsx):
        records = load_individu_survey(temp_survey_indiv_xlsx)
        assert len(records) == 1
        assert "NIK" in records[0]
        assert "_source_file" in records[0]

    def test_load_keluarga(self, temp_dir, keluarga_survey):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [k for k in keluarga_survey if not k.startswith("_")]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c).value = h
        for c, h in enumerate(headers, 1):
            ws.cell(2, c).value = keluarga_survey.get(h)
        fpath = os.path.join(temp_dir, "test_keluarga.xlsx")
        wb.save(fpath)

        records = load_keluarga_survey(fpath)
        assert len(records) == 1
        assert "Nomor KK" in records[0]

    def test_load_skips_empty_rows(self, temp_dir, keluarga_survey):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [k for k in keluarga_survey if not k.startswith("_")][:3]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c).value = h
        # Row 2 has data, row 3 is fully empty — should be skipped
        ws.cell(2, 1).value = "3505201010060507"
        fpath = os.path.join(temp_dir, "test_keluarga.xlsx")
        wb.save(fpath)

        records = load_keluarga_survey(fpath)
        assert len(records) == 1
