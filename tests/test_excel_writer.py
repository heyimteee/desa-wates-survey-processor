"""Tests for excel_writer.py."""

import datetime
import os

import openpyxl
import pytest

from excel_writer import write_individu_output, write_keluarga_output
from logger import IssueLogger


class TestWriteIndividuOutput:
    def test_creates_file(self, temp_dir, temp_template_indiv_xlsx):
        il = IssueLogger()
        records = [{"No": 1, "Action": "Auto", "KK": "3505201010060507",
                     "NIK": "3505200107570071", "Nama": "GITO",
                     "No. Hp": "085731939503", "No. Wa": "085731939503",
                     "Email": "test@gmail.com"}]
        out = os.path.join(temp_dir, "out.xlsx")
        write_individu_output(records, temp_template_indiv_xlsx, out, il)
        assert os.path.exists(out)

    def test_text_columns_as_string(self, temp_dir, temp_template_indiv_xlsx):
        il = IssueLogger()
        records = [{"No": 1, "Action": "Auto", "KK": "3505201010060507",
                     "NIK": "3505200107570071", "Nama": "GITO"}]
        out = os.path.join(temp_dir, "out.xlsx")
        write_individu_output(records, temp_template_indiv_xlsx, out, il)

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 4).number_format == "@"  # NIK col

    def test_date_columns_formatted(self, temp_dir, temp_template_indiv_xlsx):
        il = IssueLogger()
        d = datetime.date(1971, 7, 7)
        records = [{"No": 1, "Action": "Auto", "Tanggal_lahir": d}]
        out = os.path.join(temp_dir, "out.xlsx")
        write_individu_output(records, temp_template_indiv_xlsx, out, il)

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 10).number_format == "dd-mm-yyyy"

    def test_empty_values_become_dash(self, temp_dir, temp_template_indiv_xlsx):
        il = IssueLogger()
        records = [{"No": 1, "Action": "Auto"}]
        out = os.path.join(temp_dir, "out.xlsx")
        write_individu_output(records, temp_template_indiv_xlsx, out, il)

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        assert ws.cell(2, 4).value == "-"  # NIK

    def test_internal_keys_skipped(self, temp_dir, temp_template_indiv_xlsx):
        il = IssueLogger()
        records = [{"No": 1, "Action": "Auto", "_nik": "secret"}]
        out = os.path.join(temp_dir, "out.xlsx")
        write_individu_output(records, temp_template_indiv_xlsx, out, il)
        # Should not crash on _nik

    def test_logs_output_info(self, temp_dir, temp_template_indiv_xlsx):
        il = IssueLogger()
        records = [{"No": 1, "Action": "Auto"}]
        out = os.path.join(temp_dir, "out.xlsx")
        write_individu_output(records, temp_template_indiv_xlsx, out, il)
        assert any("Wrote 1 rows" in i["Description"] for i in il.issues)
