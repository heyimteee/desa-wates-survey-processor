"""Tests for bip_loader.py."""

import os

import openpyxl
import pytest

from bip_loader import load_bip_pool
from logger import IssueLogger


class TestLoadBipPool:
    def test_loads_single_file(self, temp_dir, sample_bip_head):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["NO", "NO KK", "NAMA LENGKAP", "NIK", "JK",
                   "TEMPAT LAHIR", "TANGGAL LAHIR", "GOL. DRH", "AGAMA",
                   "STATUS", "SHDK", "PENDIDIKAN", "PEKERJAAN",
                   "NAMA IBU", "NAMA AYAH", "ALAMAT", "NO RT", "NO RW"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c).value = h
        for c, h in enumerate(headers, 1):
            ws.cell(2, c).value = sample_bip_head.get(h)
        fpath = os.path.join(temp_dir, "test.xlsx")
        wb.save(fpath)

        by_nik, by_kk, bip_all, logger, by_name = load_bip_pool(temp_dir)
        assert len(bip_all) == 1
        assert sample_bip_head["NIK"] in by_nik

    def test_empty_folder(self, temp_dir):
        by_nik, by_kk, bip_all, logger, by_name = load_bip_pool(temp_dir)
        assert len(bip_all) == 0

    def test_returns_five_tuple(self, temp_dir, sample_bip_head):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "NO"
        ws.cell(1, 4).value = "NIK"
        ws.cell(1, 2).value = "NO KK"
        ws.cell(1, 3).value = "NAMA LENGKAP"
        ws.cell(2, 1).value = "1"
        ws.cell(2, 4).value = "3505200107570071"
        ws.cell(2, 2).value = "3505201010060507"
        ws.cell(2, 3).value = "GITO"
        fpath = os.path.join(temp_dir, "test.xlsx")
        wb.save(fpath)

        result = load_bip_pool(temp_dir)
        assert len(result) == 5
        by_nik, by_kk, bip_all, logger, by_name = result
        assert isinstance(by_nik, dict)
        assert isinstance(by_kk, dict)
        assert isinstance(bip_all, list)
        assert isinstance(by_name, dict)

    def test_kk_indexing(self, temp_dir, sample_bip_head):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["NO", "NO KK", "NAMA LENGKAP", "NIK"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c).value = h
        ws.cell(2, 1).value = "1"
        ws.cell(2, 2).value = "3505201010060507"
        ws.cell(2, 3).value = "GITO"
        ws.cell(2, 4).value = "3505200107570071"
        fpath = os.path.join(temp_dir, "test.xlsx")
        wb.save(fpath)

        by_nik, by_kk, bip_all, logger, by_name = load_bip_pool(temp_dir)
        assert "3505201010060507" in by_kk
        assert len(by_kk["3505201010060507"]) == 1

    def test_skips_temp_files(self, temp_dir):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(1, 1).value = "NO"
        ws.cell(1, 4).value = "NIK"
        ws.cell(2, 1).value = "1"
        ws.cell(2, 4).value = "3505200107570071"
        temp_path = os.path.join(temp_dir, "~$temp.xlsx")
        wb.save(temp_path)

        by_nik, by_kk, bip_all, logger, by_name = load_bip_pool(temp_dir)
        assert len(bip_all) == 0

    def test_name_index_populated(self, temp_dir, sample_bip_head):
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["NO", "NO KK", "NAMA LENGKAP", "NIK", "NO RT", "NO RW"]
        for c, h in enumerate(headers, 1):
            ws.cell(1, c).value = h
        ws.cell(2, 1).value = "1"
        ws.cell(2, 2).value = "3505201010060507"
        ws.cell(2, 3).value = "GITO"
        ws.cell(2, 4).value = "3505200107570071"
        ws.cell(2, 5).value = "005"
        ws.cell(2, 6).value = "001"
        fpath = os.path.join(temp_dir, "test.xlsx")
        wb.save(fpath)

        by_nik, by_kk, bip_all, logger, by_name = load_bip_pool(temp_dir)
        assert "gito" in by_name
        assert len(by_name["gito"]) == 1
