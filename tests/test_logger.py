"""Tests for logger.py."""

import os

import openpyxl
import pytest

from logger import IssueLogger


class TestIssueLogger:
    def test_log_stores_issue(self):
        il = IssueLogger()
        il.log("TEST", "WARNING", "test issue", nik="123", nama="Test")
        assert len(il.issues) == 1
        assert il.issues[0]["Category"] == "TEST"
        assert il.issues[0]["Severity"] == "WARNING"

    def test_log_validation_stores(self):
        il = IssueLogger()
        il.log_validation("1", "KK1", "NIK1", "NAME", "Data Valid", "note")
        assert len(il.validations) == 1
        assert il.validations[0]["No"] == "1"

    def test_write_creates_file(self, temp_dir):
        il = IssueLogger()
        il.log("TEST", "INFO", "test")
        fpath = os.path.join(temp_dir, "test_issues.xlsx")
        il.write_to_excel(fpath)
        assert os.path.exists(fpath)

    def test_write_creates_data_issues_sheet(self, temp_dir):
        il = IssueLogger()
        il.log("TEST", "INFO", "test")
        fpath = os.path.join(temp_dir, "test_issues.xlsx")
        il.write_to_excel(fpath)
        wb = openpyxl.load_workbook(fpath)
        assert "DATA ISSUES" in wb.sheetnames

    def test_write_with_validations_creates_second_sheet(self, temp_dir):
        il = IssueLogger()
        il.log("TEST", "INFO", "test")
        il.log_validation("1", "KK1", "NIK1", "NAME", "Valid", "ok")
        fpath = os.path.join(temp_dir, "test_issues.xlsx")
        il.write_to_excel(fpath)
        wb = openpyxl.load_workbook(fpath)
        assert "DATA VALIDASI" in wb.sheetnames

    def test_no_validations_single_sheet(self, temp_dir):
        il = IssueLogger()
        il.log("TEST", "INFO", "test")
        fpath = os.path.join(temp_dir, "test_issues.xlsx")
        il.write_to_excel(fpath)
        wb = openpyxl.load_workbook(fpath)
        assert len(wb.sheetnames) == 1

    def test_validation_headers(self, temp_dir):
        il = IssueLogger()
        il.log_validation("1", "KK1", "NIK1", "NAME", "Valid", "ok")
        fpath = os.path.join(temp_dir, "test_issues.xlsx")
        il.write_to_excel(fpath)
        wb = openpyxl.load_workbook(fpath)
        ws = wb["DATA VALIDASI"]
        headers = [ws.cell(1, c).value for c in range(1, 7)]
        assert headers == ["No", "KK", "NIK", "Nama", "Kategori", "Catatan"]

    def test_issue_headers(self, temp_dir):
        il = IssueLogger()
        il.log("TEST", "INFO", "test")
        fpath = os.path.join(temp_dir, "test_issues.xlsx")
        il.write_to_excel(fpath)
        wb = openpyxl.load_workbook(fpath)
        ws = wb["DATA ISSUES"]
        headers = [ws.cell(1, c).value for c in range(1, 7)]
        assert headers == ["Category", "Severity", "NIK", "Nama", "KK", "Description"]
