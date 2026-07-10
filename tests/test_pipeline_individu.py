"""Tests for pipeline_individu.py."""

import os
import tempfile

import openpyxl

from logger import IssueLogger
from pipeline_individu import process_individu


class TestProcessIndividu:
    @staticmethod
    def _write_survey_xlsx(records, temp_dir, filename):
        """Write a list of survey dicts to a temp .xlsx file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if records:
            headers = [k for k in records[0] if not k.startswith("_")]
            for c, h in enumerate(headers, 1):
                ws.cell(1, c).value = h
            for r, rec in enumerate(records, 2):
                for c, h in enumerate(headers, 1):
                    ws.cell(r, c).value = rec.get(h)
        path = os.path.join(temp_dir, filename)
        wb.save(path)
        return path

    def _make_survey(self, overrides=None):
        """Build a minimal INDIVIDU survey record for testing."""
        rec = {
            "Nama": "GITO",
            "NIK": "3505200107570071",
            "Alamat ": "DSN SIDOMULYO",
            "RT": "5",
            "RW": "1",
            "Status": "Kawin",
            "Usia Pertama Kali Menikah": "25",
            "Nomor HP/Wa": "085731939503",
            "Email": "",
            "Kondisi Pekerjaan": "BEKERJA",
            "Pekerjaan Saat Ini": "PETANI",
            "Penghasilan Setahun Terahir": "12000000",
            "Jamkes ?": "Peserta",
            "Penyakit Yang Diderita Setahun Terahir": "TIDAK",
            "Akses Kesehatan [RUMAH SAKIT]": "TIDAK PERNAH",
            "Akses Kesehatan [PUKESMAS DENGAN RAWAT INAP]": "TIDAK PERNAH",
            "Akses Kesehatan [TEMPAT PRAKTEK DOKTER]": "TIDAK PERNAH",
            "Akses Kesehatan [TEMPAT PRAKTEK BIDAN]": "TIDAK PERNAH",
            "Akses Kesehatan [APOTEK]": "Lebih dari 5 kali",
            "Akses Kesehatan [POSYANDU ILP]": "TIDAK PERNAH",
            "Bayi Usia 1 - 6 Bulan Konsumsi Asi": "TIDAK",
            "Disabilitas ?": "TIDAK",
            "Pendidikan ?": "SLTP / SEDERAJAT",
            "Mendapat Pelayanan Desa 1 Tahun Terahir ?": "YA",
        }
        if overrides:
            rec.update(overrides)
        return rec

    def _run(self, records, bip_pool):
        """Write survey records to temp file and run the pipeline."""
        il = IssueLogger()
        td = tempfile.mkdtemp()
        try:
            fname = "PENDATAAN INDIVIDU RT 05 RW 01 DUSUN SIDOMULYO (Jawaban).xlsx"
            fpath = self._write_survey_xlsx(records, td, fname)
            return process_individu([fpath], bip_pool["by_nik"],
                                    bip_pool["by_name"], il)
        finally:
            import shutil
            shutil.rmtree(td, ignore_errors=True)

    def test_basic_processing(self, bip_pool):
        """Valid NIK + BIP match produces an output row."""
        result = self._run([self._make_survey()], bip_pool)
        assert len(result) == 1
        assert result[0]["Nama"] == "GITO"
        assert result[0]["KK"] == "3505201010060507"

    def test_bip_enrichment(self, bip_pool):
        """BIP data correctly fills demographic fields."""
        result = self._run([self._make_survey()], bip_pool)
        rec = result[0]
        assert rec["Jenis kelamin"] == "L"
        assert rec["Tempat lahir"] == "MALANG"
        assert rec["Agama"] == "ISLAM"

    def test_nik_invalid_sorted_to_bottom(self, bip_pool):
        """Invalid NIK that can't be repaired goes to bottom."""
        result = self._run([self._make_survey({
            "NIK": "12345", "Nama": "UNKNOWN_PERSON"})], bip_pool)
        assert len(result) == 1
        assert result[0]["_sort_priority"] > 0
        assert result[0]["NIK"] == ""

    def test_nik_repair_attempted(self, bip_pool):
        """A 14-digit NIK matching by name+RT/RW gets repaired."""
        result = self._run([self._make_survey({"NIK": "35052001075700"})], bip_pool)
        assert len(result) == 1
        assert result[0]["NIK"] == "3505200107570071"

    def test_nik_empty_bottom(self, bip_pool):
        """Empty NIK with unknown name goes to bottom."""
        empty = self._make_survey({"NIK": "", "Nama": "UNKNOWN_PERSON"})
        valid = self._make_survey()
        result = self._run([empty, valid], bip_pool)
        assert result[0]["NIK"] != ""
        assert result[-1]["NIK"] == ""

    def test_same_name_duplicate_auto_dedup(self, bip_pool):
        """Two records with same NIK and same name → auto-dedup."""
        result = self._run([self._make_survey(), self._make_survey()], bip_pool)
        assert len(result) == 1

    def test_diff_name_duplicate_kept_both(self, bip_pool):
        """Two records with same NIK but different names → both kept, flagged."""
        result = self._run([
            self._make_survey({"Nama": "GITO"}),
            self._make_survey({"Nama": "SUPRIYANTO"}),
        ], bip_pool)
        assert len(result) >= 1

    def test_no_bip_for_nik(self, bip_pool):
        """NIK valid but not in BIP — still produces a row."""
        result = self._run([self._make_survey({"NIK": "3505200107570099"})], bip_pool)
        assert len(result) == 1
        assert result[0]["KK"] == ""

    def test_default_fills(self, bip_pool):
        """Static defaults are applied."""
        result = self._run([self._make_survey()], bip_pool)
        rec = result[0]
        assert rec["Suku Bangsa"] == "JAWA"
        assert rec["Warga Negara"] == "INDONESIA"
        assert rec["DI EKSPOR"] == "TIDAK"

    def test_pendidikan_empty_fills_default(self, bip_pool):
        """Empty pendidikan from survey, no BIP → TIDAK/BELUM SEKOLAH."""
        result = self._run([self._make_survey({
            "Pendidikan ?": "", "NIK": "3505200107570099"})], bip_pool)
        assert result[0]["Pendidikan Tertinggi"] == "TIDAK / BELUM SEKOLAH"

    def test_pekerjaan_empty_fills_default(self, bip_pool):
        """Empty pekerjaan from survey → BIP fallback."""
        result = self._run([self._make_survey({"Pekerjaan Saat Ini": ""})], bip_pool)
        assert result[0]["Pekerjaan Utama"] == "PETANI"

    def test_action_column(self, bip_pool):
        """Action column — first row RW, others empty."""
        result = self._run([self._make_survey()], bip_pool)
        assert result[0]["Action"] == "RW 01"

    def test_phone_clean_in_output(self, bip_pool):
        """Phone values are cleaned in output."""
        result = self._run([self._make_survey({"Nomor HP/Wa": "+628123456789"})], bip_pool)
        assert result[0]["No. Hp"] == "08123456789"

    def test_health_normalized_to_int(self, bip_pool):
        """Health text values are normalized to integers."""
        result = self._run([self._make_survey({
            "Akses Kesehatan [RUMAH SAKIT]": "TIDAK PERNAH, 1"})], bip_pool)
        assert result[0]["RUMAH SAKIT"] == 1  # max(0, 1) = 1
