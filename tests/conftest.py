"""Shared test fixtures and factories."""

import datetime
import os
import tempfile

import openpyxl
import pytest
from openpyxl.styles import Font

from utils import build_bip_name_index


@pytest.fixture
def sample_bip_head():
    """Realistic BIP record for a kepala keluarga."""
    return {
        "NO": "1",
        "NO KK": "3505201010060507",
        "NAMA LENGKAP": "GITO",
        "NIK": "3505200107570071",
        "JK": "L",
        "TEMPAT LAHIR": "MALANG",
        "TANGGAL LAHIR": "07-07-1971",
        "GOL. DRH": "A",
        "AGAMA": "ISLAM",
        "STATUS": "KAWIN",
        "SHDK": "KEPALA KELUARGA",
        "PENDIDIKAN": "SLTP/SEDERAJAT",
        "PEKERJAAN": "PETANI",
        "NAMA IBU": "SRIYATUN",
        "NAMA AYAH": "SUKARMAN",
        "ALAMAT": "DSN SIDOMULYO",
        "NO RT": "005",
        "NO RW": "001",
    }


@pytest.fixture
def sample_bip_spouse():
    """Realistic BIP record for an istri."""
    return {
        "NO": "2",
        "NO KK": "3505201010060507",
        "NAMA LENGKAP": "SUMINEM",
        "NIK": "3505204107750001",
        "JK": "P",
        "TEMPAT LAHIR": "MALANG",
        "TANGGAL LAHIR": "15-07-1975",
        "GOL. DRH": "B",
        "AGAMA": "ISLAM",
        "STATUS": "KAWIN",
        "SHDK": "ISTRI",
        "PENDIDIKAN": "SD/SEDERAJAT",
        "PEKERJAAN": "TIDAK BEKERJA",
        "NAMA IBU": "",
        "NAMA AYAH": "",
        "ALAMAT": "DSN SIDOMULYO",
        "NO RT": "005",
        "NO RW": "001",
    }


@pytest.fixture
def bip_pool(sample_bip_head, sample_bip_spouse):
    """Full BIP pool: by_nik, by_kk, all_records, by_name."""
    all_records = [sample_bip_head, sample_bip_spouse]
    by_nik = {r["NIK"]: r for r in all_records}
    by_kk = {}
    for r in all_records:
        by_kk.setdefault(r["NO KK"], []).append(r)
    by_name = build_bip_name_index(all_records)
    return {
        "by_nik": by_nik,
        "by_kk": by_kk,
        "all": all_records,
        "by_name": by_name,
    }


@pytest.fixture
def indiv_survey():
    """Realistic INDIVIDU survey record."""
    return {
        "Timestamp": datetime.datetime(2025, 10, 18, 12, 5, 30),
        "Nama": "GITO",
        "NIK": "3505200107570071",
        "Alamat ": "DSN SIDOMULYO",
        "RT": "5",
        "RW": "1",
        "Status": "Kawin",
        "Usia Pertama Kali Menikah": "25",
        "Nomor HP/Wa": "085731939503",
        "Email": "",
        "Kondisi Pekerjaan": "BEKERJA",
        "Pekerjaan Saat Ini": "PETANI",
        "Penghasilan Setahun Terahir": "12000000",
        "Jamkes ?": "Peserta",
        "Penyakit Yang Diderita Setahun Terahir": "TIDAK",
        "Akses Kesehatan [RUMAH SAKIT]": "TIDAK PERNAH",
        "Akses Kesehatan [PUKESMAS DENGAN RAWAT INAP]": "TIDAK PERNAH",
        "Akses Kesehatan [TEMPAT PRAKTEK DOKTER]": 2.0,
        "Akses Kesehatan [TEMPAT PRAKTEK BIDAN]": 3.0,
        "Akses Kesehatan [APOTEK]": "Lebih dari 5 kali",
        "Akses Kesehatan [POSYANDU ILP]": "TIDAK PERNAH",
        "Bayi Usia 1 - 6 Bulan Konsumsi Asi": "TIDAK",
        "Disabilitas ?": "TIDAK",
        "Pendidikan ?": "SLTP / SEDERAJAT",
        "Mendapat Pelayanan Desa 1 Tahun Terahir ?": "YA",
        "_source_file": "PENDATAAN DATA MIKRO INDIVIDU RT 05 RW 01 DUSUN SIDOMULYO (Jawaban).xlsx",
    }


@pytest.fixture
def keluarga_survey():
    """Realistic KELUARGA survey record."""
    return {
        "Timestamp": datetime.datetime(2025, 10, 18, 12, 5, 30),
        "Nomor KK": "3505201010060507",
        "Nomor NIK Kepala Keluarga": "3505200107570071",
        "Nama": "GITO",
        "RT": "5",
        "RW": "1",
        "Tempat Tinggal Yang Ditempati": "Milik Sendiri",
        "Luas Tempat Tinggal ?": 72,
        "Jenis Lantai tempat tinggal": "Keramik",
        "Jenis Dinding Rumah": "Semen/Beton/Kayu berkualitas tinggi",
        "Energi Untuk Memasak": "LPG",
        "Fasilitas MCK": "Milik Sendiri",
        "Sumber Air Mandi terbanyak": "Mata air/Sumur",
        "Sumber Air Minum terbanyak": "Mata air/Sumur",
        "Jarak dari rumah ke PAUD": 100,
        "Waktu yang dibutuhkan dari rumah ke PAUD": 5,
        "Jarak dari rumah ke TK/RA": 100,
        "Waktu yang dibutuhkan dari rumah ke TK/RA": 5,
        "Jarak dari rumah ke SD/MI atau Sederajat": 200,
        "Waktu yang dibutuhkan dari rumah ke SD/MI atau Sederajat": 8,
        "Jarak dari rumah ke SMP/MTs atau Sederajat": 1800,
        "Waktu yang dibutuhkan dari rumah ke SMP/MTs atau Sederajat": 20,
        "Jarak dari rumah ke SMA/MA atau Sederajat": 2500,
        "Waktu yang dibutuhkan dari rumah ke  SMA/MA atau Sederajat": 30,
        "Jarak dari rumah ke Perguruan Tinggi": 18000,
        "Waktu yang dibutuhkan dari rumah ke Perguruan Tinggi": 60,
        "Jarak dari rumah ke Rumah Sakit": 35000,
        "Waktu yang dibutuhkan dari rumah ke Rumah Sakit": 45,
        "Jarak dari rumah ke Rumah Bersalin": 3000,
        "Waktu yang dibutuhkan dari rumah ke Rumah Bersalin": 15,
        "Jarak dari rumah ke Pukesmas": 1000,
        "Waktu yang dibutuhkan dari rumah ke Pukesmas": 10,
        "Jarak dari rumah ke Posyandu": 200,
        "Waktu yang dibutuhkan dari rumah ke Posyandu": 5,
        "Jarak dari rumah ke Apotek": 800,
        "Waktu yang dibutuhkan dari rumah ke  Apotek": 10,
        "Pemanfaat/Penerima Program pemerintah [BLT]": "Tidak",
        "Pemanfaat/Penerima Program pemerintah [PKH]": "Tidak",
        "Pemanfaat/Penerima Program pemerintah [BPNT]": "Tidak",
        "Pemanfaat/Penerima Program pemerintah [BST]": "Tidak",
        "Pemanfaat/Penerima Program pemerintah [Sembako]": "Tidak",
        "Pemanfaat/Penerima Program pemerintah [BSU]": "Tidak",
        "Pemanfaat/Penerima Program pemerintah [Bantuan UMKM]": "Tidak",
        "Pemanfaat/Penerima Program pemerintah [Bantuan Untuk Pekerja]": "Tidak",
        "Pemanfaat/Penerima Program pemerintah [Bantuan untuk Pendidikan Anak]": "Tidak",
        "Pemanfaat/Penerima Program pemerintah [Bantuan Lain dari Pemerintah ]": "Tidak",
        "Rata rata pengeluaran satu keluarga dalam satu bulan": "500.000 sd 1.000.000",
        "_source_file": "Pendataan Data Mikro Keluarga RT 05 RW 01 Dusun Sidomulyo (Jawaban).xlsx",
    }


@pytest.fixture
def temp_dir():
    """Temporary directory that auto-cleans."""
    with tempfile.TemporaryDirectory() as d:
        yield d


@pytest.fixture
def temp_bip_xlsx(temp_dir, sample_bip_head):
    """A real temp BIP .xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["NO", "NO KK", "NAMA LENGKAP", "NIK", "JK", "TEMPAT LAHIR",
               "TANGGAL LAHIR", "GOL. DRH", "AGAMA", "STATUS", "SHDK",
               "PENDIDIKAN", "PEKERJAAN", "NAMA IBU", "NAMA AYAH",
               "ALAMAT", "NO RT", "NO RW"]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
    for c, h in enumerate(headers, 1):
        ws.cell(2, c).value = sample_bip_head.get(h)
    path = os.path.join(temp_dir, "test_bip.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def temp_survey_indiv_xlsx(temp_dir, indiv_survey):
    """A real temp INDIVIDU survey .xlsx file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(indiv_survey.keys())
    headers = [h for h in headers if not h.startswith("_")]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
    for c, h in enumerate(headers, 1):
        ws.cell(2, c).value = indiv_survey.get(h)
    path = os.path.join(temp_dir, "test_indiv_survey.xlsx")
    wb.save(path)
    return path


@pytest.fixture
def temp_template_indiv_xlsx(temp_dir):
    """A minimal INDIVIDU template for writing tests."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DATA INDIVIDU"
    headers = ["Action", "No", "KK", "NIK", "Gelar awal", "Nama",
               "Gelar akhir", "Jenis kelamin", "Tempat lahir",
               "Tanggal_lahir", "Usia", "Status",
               "Usia Saat pertama kali menikah", "Agama",
               "Suku Bangsa", "Warga Negara", "No. Hp", "No. Wa",
               "Email", "Facebook", "Twitter", "Instagram",
               "Kondisi pekerjaan", "Pekerjaan Utama",
               "Jaminan Sosial Ketenagakerjaan",
               "Penghasilan Setahun Terakhir", "SUMBER PENGHASILAN",
               "JUMLAH ASET DARI SUMBER PENGHASILAN", "SATUAN",
               "PENGHASILAN SETAHUN", "DI EKSPOR",
               "PENYAKIT YANG DIDERITA SETAHUN TERAKHIR",
               "RUMAH SAKIT", "RUMAH SAKIT BERSALIN",
               "PUSKESMAS DENGAN RAWAT INAP",
               "PUSKESMAS TANPA RAWAT INAP", "PUSKEMAS PEMBANTU",
               "POLIKLINIK", "TEMPAT PRAKTEK DOKTER", "RUMAH BERSALIN",
               "TEMPAT PRAKTEK BIDAN", "POSKESDES", "POLINDES",
               "APOTIK", "TOKO KHUSUS OBAT / JAMU", "POSYANDU",
               "POSBINDU", "TEMPAT PRAKTIK DUKUN BAYI / BERSALIN",
               "JAMKES", "BAYI Usia 1-6 bulan Konsumsi ASI",
               "JENIS DISABILITAS", "Pendidikan Tertinggi",
               "Berapa Tahun mengenyam pendidikan dasar (SD,SMP,SMA)",
               "Pendidikan yang sedang di ikuti",
               "Bahasa yang digunakan di Rumah dan Pemukiman",
               "Bahasa yang digunakan di Lembaga Formal",
               "Jumlah kerja bakti 1 tahun terakhir",
               "Siskamling 1 tahun terakhir",
               "Pesta Rakyat (Adat) 1 tahun terakhir",
               "Frekuensi Melayat 1 tahun terakhir",
               "Frekuensi besuk orang sakit 1 tahun terakhir",
               "Frekuensi menolong kecelakaan 1 tahun terakhir",
               "Mendapatkan Pelayanan Desa 1 tahun terakhir",
               "Bagaimana pelayanan desa yang diperoleh?",
               "Pernah menyampaikan masukan/saran kepada pihak Desa?",
               "Bagaimana keterbukaan desa terhadap masukan?",
               "Terjadi Bencana 1 tahun terakhir",
               "Apakah anda terkena dampak bencana",
               "Apakah menerima pemenuhan Kebutuhan Dasar saat Bencana (makanan,pakaian, tempat tinggal)?",
               "Apakah ada penanganan psikososial keluarga terdampak bencana (penyuluhan/konseling/terapi)?",
               ]
    for c, h in enumerate(headers, 1):
        ws.cell(1, c).value = h
    path = os.path.join(temp_dir, "TEMPLATE_DATA_INDIVIDU.xlsx")
    wb.save(path)
    return path
