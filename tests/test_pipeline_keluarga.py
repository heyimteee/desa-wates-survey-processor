"""Tests for pipeline_keluarga.py."""

import os
import shutil
import tempfile

import openpyxl

from logger import IssueLogger
from pipeline_keluarga import process_keluarga, _clean_distance, _clean_time


class TestCleanDistance:
    def test_meters_converted_to_km(self):
        """1000 raw meters → 1 km."""
        km, note = _clean_distance(1000, 10, "TEST")
        assert km == 1.0
        assert note == ""

    def test_75_meters(self):
        """75 raw meters → 0.075 km."""
        km, note = _clean_distance(75, 2, "TEST")
        assert km == 0.075

    def test_km_entry_detected(self):
        """raw=1 with time=5min → detected as KM entry (not 0.001km)."""
        km, note = _clean_distance(1, 5, "TEST")
        assert km == 1.0
        assert "kemungkinan sudah dalam KM" in note

    def test_extreme_returns_dash(self):
        """1,000,000,000 raw → '-'."""
        km, note = _clean_distance(1_000_000_000, 60, "TEST")
        assert km == "-"
        assert "ekstrem" in note

    def test_far_but_not_extreme(self):
        """150,000 raw (150km) → kept with warning."""
        km, note = _clean_distance(150_000, 60, "TEST")
        assert km == 150.0
        assert "jauh" in note

    def test_meters_ge_100_assumed_meters(self):
        """raw=150 (150 meters) with time 3 min → 0.15 km."""
        km, note = _clean_distance(150, 3, "TEST")
        assert km == 0.15
        assert note == ""

    def test_none_returns_dash(self):
        km, note = _clean_distance(None, None, "TEST")
        assert km == "-"

    def test_text_value_preserved(self):
        km, note = _clean_distance("unknown", None, "TEST")
        assert km == "unknown"

    def test_small_value_without_time(self):
        """Small value without time data → meters assumption."""
        km, note = _clean_distance(10, None, "TEST")
        assert km == 0.01


class TestCleanTime:
    def test_normal_time(self):
        """5 minutes → 0.083 hours, no anomaly."""
        jam, note = _clean_time(0.1, 0.083, "TEST")
        assert jam == 0.083
        assert note == ""

    def test_extreme_time(self):
        """800 minutes (13.3 hours) → '-'."""
        jam, note = _clean_time(10, 13.3, "TEST")
        assert jam == "-"
        assert "ekstrem" in note

    def test_absurd_speed(self):
        """0.001 km in 5 hours → speed 0.0002 km/h → '-'."""
        jam, note = _clean_time(0.001, 5.0, "TEST")
        assert jam == "-"
        assert "kecepatan tidak wajar" in note

    def test_non_numeric_preserved(self):
        jam, note = _clean_time(0.1, "unknown", "TEST")
        assert jam == "unknown"


class TestProcessKeluarga:
    @staticmethod
    def _write_survey_xlsx(records, temp_dir, filename):
        """Write a list of survey dicts to a temp .xlsx file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        if records:
            headers = [k for k in records[0] if not k.startswith("_")]
            for c, h in enumerate(headers, 1):
                ws.cell(1, c).value = h
            for r, rec in enumerate(records, 2):
                for c, h in enumerate(headers, 1):
                    ws.cell(r, c).value = rec.get(h)
        path = os.path.join(temp_dir, filename)
        wb.save(path)
        return path

    def _make_survey(self, overrides=None):
        """Build a minimal KELUARGA survey record."""
        rec = {
            "Nomor KK": "3505201010060507",
            "Nomor NIK Kepala Keluarga": "3505200107570071",
            "Nama": "GITO",
            "RT": "5",
            "RW": "1",
            "Tempat Tinggal Yang Ditempati": "Milik Sendiri",
            "Luas Tempat Tinggal ?": 72,
            "Jenis Lantai tempat tinggal": "Keramik",
            "Jenis Dinding Rumah": "Semen/Beton/Kayu berkualitas tinggi",
            "Energi Untuk Memasak": "LPG",
            "Fasilitas MCK": "Milik Sendiri",
            "Sumber Air Mandi terbanyak": "Mata air/Sumur",
            "Sumber Air Minum terbanyak": "Mata air/Sumur",
            "Jarak dari rumah ke PAUD": 100,
            "Waktu yang dibutuhkan dari rumah ke PAUD": 5,
            "Jarak dari rumah ke TK/RA": 100,
            "Waktu yang dibutuhkan dari rumah ke TK/RA": 5,
            "Jarak dari rumah ke SD/MI atau Sederajat": 200,
            "Waktu yang dibutuhkan dari rumah ke SD/MI atau Sederajat": 8,
            "Jarak dari rumah ke SMP/MTs atau Sederajat": 1800,
            "Waktu yang dibutuhkan dari rumah ke SMP/MTs atau Sederajat": 20,
            "Jarak dari rumah ke SMA/MA atau Sederajat": 2500,
            "Waktu yang dibutuhkan dari rumah ke  SMA/MA atau Sederajat": 30,
            "Jarak dari rumah ke Perguruan Tinggi": 18000,
            "Waktu yang dibutuhkan dari rumah ke Perguruan Tinggi": 60,
            "Jarak dari rumah ke Rumah Sakit": 35000,
            "Waktu yang dibutuhkan dari rumah ke Rumah Sakit": 45,
            "Jarak dari rumah ke Rumah Bersalin": 3000,
            "Waktu yang dibutuhkan dari rumah ke Rumah Bersalin": 15,
            "Jarak dari rumah ke Pukesmas": 1000,
            "Waktu yang dibutuhkan dari rumah ke Pukesmas": 10,
            "Jarak dari rumah ke Posyandu": 200,
            "Waktu yang dibutuhkan dari rumah ke Posyandu": 5,
            "Jarak dari rumah ke Apotek": 800,
            "Waktu yang dibutuhkan dari rumah ke  Apotek": 10,
            "Pemanfaat/Penerima Program pemerintah [BLT]": "Tidak",
            "Pemanfaat/Penerima Program pemerintah [PKH]": "Tidak",
            "Pemanfaat/Penerima Program pemerintah [BPNT]": "Tidak",
            "Pemanfaat/Penerima Program pemerintah [BST]": "Tidak",
            "Pemanfaat/Penerima Program pemerintah [Sembako]": "Tidak",
            "Pemanfaat/Penerima Program pemerintah [BSU]": "Tidak",
            "Pemanfaat/Penerima Program pemerintah [Bantuan UMKM]": "Tidak",
            "Pemanfaat/Penerima Program pemerintah [Bantuan Untuk Pekerja]": "Tidak",
            "Pemanfaat/Penerima Program pemerintah [Bantuan untuk Pendidikan Anak]": "Tidak",
            "Pemanfaat/Penerima Program pemerintah [Bantuan Lain dari Pemerintah ]": "Tidak",
            "Rata rata pengeluaran satu keluarga dalam satu bulan": "500.000 sd 1.000.000",
        }
        if overrides:
            rec.update(overrides)
        return rec

    def _run(self, records, bip_pool):
        """Write survey records to temp file and run the pipeline."""
        il = IssueLogger()
        td = tempfile.mkdtemp()
        try:
            fname = "Pendataan Keluarga RT 05 RW 01 Dusun Sidomulyo (Jawaban).xlsx"
            fpath = self._write_survey_xlsx(records, td, fname)
            return process_keluarga([fpath], bip_pool["by_kk"],
                                    bip_pool["all"], il)
        finally:
            shutil.rmtree(td, ignore_errors=True)

    def test_basic_processing(self, bip_pool):
        """KK with 2 BIP members → 2 output rows."""
        result = self._run([self._make_survey()], bip_pool)
        assert len(result) == 2

    def test_nik_kk_repair(self, bip_pool):
        """15-digit NIK KK → repaired from BIP."""
        sv = self._make_survey({"Nomor NIK Kepala Keluarga": "350520010750071"})
        result = self._run([sv], bip_pool)
        assert result[0]["NIK Kepala Keluarga"] == "3505200107570071"

    def test_corrupt_row_filtered(self, bip_pool):
        """KK=1, NIK KK=1, Nama=1.0 → skipped."""
        sv = self._make_survey({
            "Nomor KK": "1", "Nomor NIK Kepala Keluarga": "1", "Nama": "1.0"})
        result = self._run([sv], bip_pool)
        assert len(result) == 0

    def test_kk_not_in_bip(self, bip_pool):
        """KK not in BIP → one row with survey data only."""
        sv = self._make_survey({"Nomor KK": "9999999999999999"})
        result = self._run([sv], bip_pool)
        assert len(result) == 1
        assert result[0]["NIK"] == ""

    def test_duplicate_kk_filtered(self, bip_pool):
        """Duplicate KK in survey → second removed."""
        sv = self._make_survey()
        result = self._run([sv, sv], bip_pool)
        assert len(result) == 2

    def test_distance_converted(self, bip_pool):
        """Distance values are converted from meters."""
        sv = self._make_survey({"Jarak ke PAUD": 1000, "Waktu PAUD": 10})
        result = self._run([self._make_survey()], bip_pool)
        assert result[0]["PAUD - JARAK (KM)"] == 0.1

    def test_time_converted(self, bip_pool):
        """Time values are converted from minutes."""
        result = self._run([self._make_survey()], bip_pool)
        assert result[0]["PAUD - WAKTU (JAM)"] == 5 / 60.0

    def test_kk_validation_note(self, bip_pool):
        """Invalid KK → flagged."""
        sv = self._make_survey({"Nomor KK": "123"})
        result = self._run([sv], bip_pool)
        notes = result[0].get("_notes", [])
        assert any("NO KK" in n for n in notes)

    def test_action_column(self, bip_pool):
        """Action column set on output."""
        result = self._run([self._make_survey()], bip_pool)
        assert result[0]["Action"] in ("Auto", "Periksa")

    def test_luas_validation(self, bip_pool):
        """LUAS LANTAI > 500 → flagged."""
        sv = self._make_survey({"Luas Tempat Tinggal ?": 600})
        result = self._run([sv], bip_pool)
        notes = result[0].get("_notes", [])
        assert any("Luas Lantai" in n for n in notes)

    def test_bantuan_ya_tidak_normalized(self, bip_pool):
        """Ambiguous YA,TIDAK → YA."""
        sv = self._make_survey({
            "Pemanfaat/Penerima Program pemerintah [BLT]": "Ya, Tidak"})
        result = self._run([sv], bip_pool)
        assert result[0]["BLT"] == "YA"

    def test_extreme_distance_becomes_dash(self, bip_pool):
        """Extreme distance (>200km) → '-'"""
        sv = self._make_survey({
            "Jarak dari rumah ke PAUD": 1_000_000_000,
            "Waktu yang dibutuhkan dari rumah ke PAUD": 60,
        })
        result = self._run([sv], bip_pool)
        assert result[0]["PAUD - JARAK (KM)"] == "-"

    def test_sort_kk_az(self, bip_pool):
        """Two KKs sorted A-Z by KK."""
        sv1 = self._make_survey({"Nomor KK": "3505201010060510"})
        sv2 = self._make_survey({"Nomor KK": "3505201010060507"})
        result = self._run([sv1, sv2], bip_pool)
        kks = [r["NO KK"] for r in result]
        assert kks == sorted(kks)
